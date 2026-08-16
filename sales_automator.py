import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import sys
import os

def create_automated_report():
    # 1. Initialize GUI for File Picker (Hides the main empty window)
    root = tk.Tk()
    root.withdraw()

    # 2. Prompt user to select the raw dataset
    messagebox.showinfo("Sales Automator", "Please select the raw Superstore Sales file (CSV or Excel).")
    input_path = filedialog.askopenfilename(
        title="Select Raw Data File",
        filetypes=[("CSV Files", "*.csv"), ("Excel Files", "*.xlsx")]
    )

    if not input_path:
        messagebox.showwarning("Cancelled", "No file selected. Exiting automation.")
        sys.exit()

    print("Processing data... Please wait.")

    # 3. Ingest Data (Handling potential encoding issues for standard CSVs)
    try:
        if input_path.endswith('.csv'):
            df = pd.read_csv(input_path, encoding='windows-1252')
        else:
            df = pd.read_excel(input_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read file:\n{e}")
        sys.exit()

    # 4. Data Transformation (Pandas)
    # Ensure Sales and Profit are numeric
    df['Sales'] = pd.to_numeric(df['Sales'], errors='coerce').fillna(0)
    df['Profit'] = pd.to_numeric(df['Profit'], errors='coerce').fillna(0)

    # Generate Pivot 1: Sales & Profit by Region
    region_summary = df.groupby('Region')[['Sales', 'Profit']].sum().reset_index()
    region_summary = region_summary.sort_values(by='Sales', ascending=False)

    # Generate Pivot 2: Sales by Category
    category_summary = df.groupby('Category')[['Sales']].sum().reset_index()

    # 5. Excel Generation & Formatting (OpenPyXL)
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"

    # Define Styles
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    currency_format = '"$"#,##0.00'

    # Function to append a styled dataframe to the worksheet
    def append_styled_table(dataframe, start_row, title):
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=14)
        
        # Write headers
        headers = list(dataframe.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=False), start_row + 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # Apply currency formatting to numeric columns
                if isinstance(value, (int, float)):
                    cell.number_format = currency_format

    # Write tables to the sheet
    append_styled_table(region_summary, start_row=2, title="Regional Sales & Profit")
    append_styled_table(category_summary, start_row=10, title="Sales by Product Category")

    # 6. Add an Automated Chart for Regions
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Revenue by Region"
    chart.y_axis.title = "Total Sales ($)"
    chart.x_axis.title = "Region"

    # Data range for chart (Starts at row 3, ends at row 3 + len(region_summary))
    data_ref = Reference(ws, min_col=2, min_row=2, max_row=2 + len(region_summary), max_col=2)
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=2 + len(region_summary))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    
    # Place chart next to the tables
    ws.add_chart(chart, "E2")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # 7. Save the File
    output_dir = os.path.dirname(input_path)
    output_file = os.path.join(output_dir, "Automated_Sales_Report.xlsx")
    wb.save(output_file)

    messagebox.showinfo("Success", f"Report generated successfully!\nSaved to:\n{output_file}")
    print(f"Done! File saved at {output_file}")

if __name__ == "__main__":
    create_automated_report()